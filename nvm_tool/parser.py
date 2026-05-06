from __future__ import annotations

import json
import logging
from pathlib import Path
from typing import Any, Callable, Dict, List, Mapping, Optional, Union
from xml.etree import ElementTree as ET

from .models import (
    AUTOSAR_NAMESPACE,
    NVM_BLOCK_CONTAINER_DEFINITION_REF,
    NVM_BLOCK_WRITE_PROT_DEFINITION_REF,
    NVM_MODULE_DEFINITION_REF,
    NvMBlock,
    ParsedArxmlDocument,
)
from .transformers import NvMBlockTransformer


class NvMConfigParser:
    """Parses JSON, Excel, and previous NvM ARXML inputs into structured models."""

    REQUIRED_FIELDS = {
        "block_name",
        "block_id",
        "block_size",
        "ram_block_name",
        "device",
        "block_management_type",
        "use_crc",
        "crc_type",
        "write_protection",
    }
    INPUT_TYPE_SUFFIXES = {
        "json": {".json"},
        "excel": {".xlsx", ".xlsm"},
    }

    def __init__(self, logger: Optional[logging.Logger] = None) -> None:
        self.logger = logger or logging.getLogger(self.__class__.__name__)

    def parse_input_file(self, input_type: str, input_path: Union[str, Path]) -> List[NvMBlock]:
        normalized_input_type = input_type.strip().lower()
        if normalized_input_type not in self.INPUT_TYPE_SUFFIXES:
            raise ValueError("Unsupported input type. Use 'json' or 'excel'.")

        path = self._validate_file_path(input_path, "Input")
        if path.suffix.lower() not in self.INPUT_TYPE_SUFFIXES[normalized_input_type]:
            expected_suffixes = ", ".join(sorted(self.INPUT_TYPE_SUFFIXES[normalized_input_type]))
            raise ValueError(
                f"Input file '{path}' does not match --input-type {normalized_input_type}. "
                f"Expected one of: {expected_suffixes}."
            )

        if normalized_input_type == "json":
            records = self._read_json(path)
            self.logger.debug("Read %d records from JSON file %s", len(records), path)
        else:
            records = self._read_excel(path)
            self.logger.debug("Read %d records from Excel file %s", len(records), path)

        blocks = [self._record_to_block(record, index) for index, record in enumerate(records, start=1)]
        self._validate_unique_ids(blocks, "input")
        self._validate_unique_generated_names(blocks, "input")
        self._warn_reserved_ids(blocks, str(path))

        sorted_blocks = sorted(blocks, key=lambda block: block.block_id)
        self.logger.info(
            "Parsed %d NvM block(s) from %s using %s mode.",
            len(sorted_blocks),
            path,
            normalized_input_type,
        )
        return sorted_blocks

    def parse_previous_arxml(self, input_path: Union[str, Path]) -> ParsedArxmlDocument:
        path = self._validate_file_path(input_path, "Previous ARXML")

        try:
            tree = ET.parse(path)
        except ET.ParseError as exc:
            raise ValueError(f"Invalid previous ARXML in {path}: {exc}.") from exc

        root = tree.getroot()
        namespace = self._extract_namespace(root.tag) or AUTOSAR_NAMESPACE
        module_configurations = self._find_nvm_module_configurations(root)
        if not module_configurations:
            raise ValueError(
                f"Previous ARXML does not contain an NvM ECUC-MODULE-CONFIGURATION-VALUES node: {path}"
            )
        if len(module_configurations) > 1:
            raise ValueError(
                f"Previous ARXML contains multiple NvM ECUC-MODULE-CONFIGURATION-VALUES nodes: {path}"
            )

        module_configuration = module_configurations[0]
        containers_element = self._find_direct_child(module_configuration, "CONTAINERS")
        blocks: List[NvMBlock] = []
        block_id_locations: Dict[int, str] = {}
        short_name_locations: Dict[str, str] = {}

        if containers_element is not None:
            for container in containers_element:
                if self._local_name(container.tag) != "ECUC-CONTAINER-VALUE":
                    continue
                definition_ref = self._find_direct_child(container, "DEFINITION-REF")
                if definition_ref is None:
                    continue
                if (definition_ref.text or "").strip() != NVM_BLOCK_CONTAINER_DEFINITION_REF:
                    continue

                block = self._parse_block_container(container)
                location = f"{path} -> SHORT-NAME '{block.short_name}'"

                if block.block_id in block_id_locations:
                    raise ValueError(
                        f"Duplicate block ID {block.block_id} found in previous ARXML at "
                        f"{block_id_locations[block.block_id]} and {location}."
                    )
                if block.short_name in short_name_locations:
                    raise ValueError(
                        f"Duplicate block name '{block.short_name}' found in previous ARXML at "
                        f"{short_name_locations[block.short_name]} and {location}."
                    )

                block_id_locations[block.block_id] = location
                short_name_locations[block.short_name] = location
                blocks.append(block)

        self._warn_reserved_ids(blocks, str(path))
        self.logger.info("Parsed %d existing NvM block(s) from %s.", len(blocks), path)

        return ParsedArxmlDocument(
            tree=tree,
            root=root,
            namespace=namespace,
            module_configuration=module_configuration,
            containers_element=containers_element,
            blocks=sorted(blocks, key=lambda block: block.block_id),
            block_id_locations=block_id_locations,
            short_name_locations=short_name_locations,
        )

    def _parse_block_container(self, container: ET.Element) -> NvMBlock:
        short_name_element = self._find_direct_child(container, "SHORT-NAME")
        short_name = (short_name_element.text or "").strip() if short_name_element is not None else ""
        if not short_name:
            raise ValueError("Previous ARXML contains an NvM block container without SHORT-NAME.")

        parameter_values = self._extract_parameter_values(container, short_name)
        return NvMBlockTransformer.from_arxml_container(short_name, parameter_values)

    def _extract_parameter_values(
        self,
        container: ET.Element,
        short_name: str,
    ) -> Dict[str, str]:
        parameter_values_element = self._find_direct_child(container, "PARAMETER-VALUES")
        if parameter_values_element is None:
            raise ValueError(f"Previous ARXML block '{short_name}' is missing PARAMETER-VALUES.")

        parameter_values: Dict[str, str] = {}
        for parameter in parameter_values_element:
            parameter_tag = self._local_name(parameter.tag)
            if parameter_tag not in {"ECUC-NUMERICAL-PARAM-VALUE", "ECUC-TEXTUAL-PARAM-VALUE"}:
                continue

            definition_ref = self._find_direct_child(parameter, "DEFINITION-REF")
            value_element = self._find_direct_child(parameter, "VALUE")
            if definition_ref is None or value_element is None:
                continue

            definition_ref_text = (definition_ref.text or "").strip()
            if not definition_ref_text:
                continue
            if definition_ref_text in parameter_values:
                raise ValueError(
                    f"Previous ARXML block '{short_name}' contains duplicate parameter '{definition_ref_text}'."
                )

            parameter_values[definition_ref_text] = (value_element.text or "").strip()

        write_protection = next(
            (
                element
                for element in container.iter()
                if self._local_name(element.tag) == "WRITE-PROTECTION"
            ),
            None,
        )
        if (
            write_protection is not None
            and NVM_BLOCK_WRITE_PROT_DEFINITION_REF not in parameter_values
        ):
            parameter_values[NVM_BLOCK_WRITE_PROT_DEFINITION_REF] = (
                write_protection.text or ""
            ).strip()

        return parameter_values

    def _find_nvm_module_configurations(self, root: ET.Element) -> List[ET.Element]:
        matches: List[ET.Element] = []
        for element in root.iter():
            if self._local_name(element.tag) != "ECUC-MODULE-CONFIGURATION-VALUES":
                continue
            definition_ref = self._find_direct_child(element, "DEFINITION-REF")
            if definition_ref is None:
                continue
            if (definition_ref.text or "").strip() == NVM_MODULE_DEFINITION_REF:
                matches.append(element)
        return matches

    def _record_to_block(self, record: Dict[str, Any], index: int) -> NvMBlock:
        self.logger.debug("Converting record %d to NvMBlock: %s", index, record)
        missing = sorted(self.REQUIRED_FIELDS - set(record))
        if missing:
            raise ValueError(f"Record {index} is missing required fields: {', '.join(missing)}")

        try:
            return NvMBlockTransformer.from_input_record(record)
        except Exception as exc:  # noqa: BLE001
            raise ValueError(f"Invalid NvM block in record {index}: {exc}") from exc

    def _read_json(self, path: Path) -> List[Dict[str, Any]]:
        try:
            payload = json.loads(path.read_text(encoding="utf-8"))
        except json.JSONDecodeError as exc:
            raise ValueError(
                f"Invalid JSON in {path}: {exc.msg} at line {exc.lineno}, column {exc.colno}."
            ) from exc

        if isinstance(payload, list):
            records = payload
        elif isinstance(payload, dict) and isinstance(payload.get("blocks"), list):
            records = payload["blocks"]
        else:
            raise ValueError("JSON input must be a list of blocks or an object with a 'blocks' list.")

        if not records:
            raise ValueError("Input does not contain any NvM blocks.")

        normalized_records: List[Dict[str, Any]] = []
        for index, record in enumerate(records, start=1):
            if not isinstance(record, Mapping):
                raise ValueError(f"Record {index} must be a JSON object.")
            normalized_records.append(dict(record))

        return normalized_records

    def _read_excel(self, path: Path) -> List[Dict[str, Any]]:
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise RuntimeError(
                "Excel input requires openpyxl. Install it with 'pip install openpyxl'."
            ) from exc

        workbook = load_workbook(filename=path, read_only=True, data_only=True)
        try:
            worksheet = workbook.active
            rows = worksheet.iter_rows(values_only=True)

            try:
                headers = next(rows)
            except StopIteration as exc:
                raise ValueError("Excel input is empty.") from exc

            normalized_headers = [self._normalize_header(header) for header in headers]
            missing = sorted(self.REQUIRED_FIELDS - set(normalized_headers))
            if missing:
                raise ValueError(
                    "Excel header row is missing required columns: " + ", ".join(missing)
                )

            duplicate_headers = self._find_duplicate_headers(normalized_headers)
            if duplicate_headers:
                raise ValueError(
                    "Excel header row contains duplicate columns after normalization: "
                    + ", ".join(duplicate_headers)
                )

            records: List[Dict[str, Any]] = []
            for row in rows:
                if row is None or all(cell in (None, "") for cell in row):
                    continue

                record = {
                    normalized_headers[position]: cell
                    for position, cell in enumerate(row)
                    if position < len(normalized_headers) and normalized_headers[position]
                }
                records.append(record)

            if not records:
                raise ValueError("Excel input does not contain any NvM blocks.")

            return records
        finally:
            workbook.close()

    def _validate_unique_ids(self, blocks: List[NvMBlock], source_label: str) -> None:
        seen: Dict[int, str] = {}
        for block in blocks:
            existing = seen.get(block.block_id)
            if existing is not None:
                raise ValueError(
                    f"Duplicate block ID {block.block_id} found in {source_label} for "
                    f"'{existing}' and '{block.block_name}'."
                )
            seen[block.block_id] = block.block_name

    def _validate_unique_generated_names(self, blocks: List[NvMBlock], source_label: str) -> None:
        self._validate_unique_strings(
            blocks,
            lambda block: block.block_id_macro,
            f"{source_label} generated block ID macro",
        )
        self._validate_unique_strings(
            blocks,
            lambda block: block.short_name,
            f"{source_label} AUTOSAR short name",
        )

    @staticmethod
    def _validate_unique_strings(
        blocks: List[NvMBlock],
        selector: Callable[[NvMBlock], str],
        label: str,
    ) -> None:
        seen: Dict[str, str] = {}
        for block in blocks:
            value = selector(block)
            existing = seen.get(value)
            if existing is not None:
                raise ValueError(
                    f"Block names '{existing}' and '{block.block_name}' collide on the {label} "
                    f"'{value}'. Rename one of the blocks."
                )
            seen[value] = block.block_name

    def _warn_reserved_ids(self, blocks: List[NvMBlock], source_label: str) -> None:
        for block in blocks:
            if block.block_id < 2:
                self.logger.warning(
                    "Block '%s' from %s uses block_id=%s. AUTOSAR typically reserves 0 and 1.",
                    block.block_name,
                    source_label,
                    block.block_id,
                )

    @staticmethod
    def _validate_file_path(input_path: Union[str, Path], label: str) -> Path:
        path = Path(input_path)
        if not path.exists():
            raise FileNotFoundError(f"{label} file not found: {path}")
        if path.is_dir():
            raise ValueError(f"{label} path must be a file, not a directory: {path}")
        return path

    @staticmethod
    def _find_duplicate_headers(headers: List[str]) -> List[str]:
        seen = set()
        duplicates = set()
        for header in headers:
            if not header:
                continue
            if header in seen:
                duplicates.add(header)
            seen.add(header)
        return sorted(duplicates)

    @staticmethod
    def _normalize_header(header: Any) -> str:
        if header is None:
            return ""
        return str(header).strip().lower().replace(" ", "_").replace("-", "_")

    @staticmethod
    def _find_direct_child(parent: ET.Element, local_name: str) -> Optional[ET.Element]:
        for child in parent:
            if NvMConfigParser._local_name(child.tag) == local_name:
                return child
        return None

    @staticmethod
    def _local_name(tag: str) -> str:
        return tag.split("}", 1)[-1]

    @staticmethod
    def _extract_namespace(tag: str) -> str:
        if tag.startswith("{") and "}" in tag:
            return tag[1:].split("}", 1)[0]
        return ""
